"""
报表生成子智能体

支持将最近一次问答结果导出为 Excel / PDF。
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib import colors


class ReportAgent:
    """导出最近一次查询/分析结果。"""

    def __init__(self, export_dir: str = "./exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)

        try:
            pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
            self.pdf_font_name = "STSong-Light"
        except Exception:
            self.pdf_font_name = "Helvetica"

    def _sanitize_filename(self, text: str, limit: int = 32) -> str:
        cleaned = re.sub(r"[^\w\u4e00-\u9fff-]+", "_", text.strip())
        cleaned = cleaned.strip("_") or "report"
        return cleaned[:limit]

    def _parse_sql_rows(self, sql_result: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if not isinstance(sql_result, dict):
            return []
        data = sql_result.get("data")
        if isinstance(data, list):
            return [row for row in data if isinstance(row, dict)]
        if isinstance(data, str):
            try:
                parsed = json.loads(data)
            except Exception:
                return []
            if isinstance(parsed, list):
                return [row for row in parsed if isinstance(row, dict)]
        return []

    def _build_summary_rows(self, payload: Dict[str, Any]) -> List[List[str]]:
        return [
            ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["问题", str(payload.get("question", ""))],
            ["意图", str(payload.get("intent", ""))],
            ["回答", str(payload.get("answer", ""))],
            ["分析摘要", str(payload.get("analysis_summary", ""))],
            ["异常摘要", str(payload.get("anomaly_summary", ""))],
            ["来源数量", str(len(payload.get("sources", []) or []))],
        ]

    def export_excel(self, payload: Dict[str, Any]) -> Path:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        for row in self._build_summary_rows(payload):
            summary_sheet.append(row)

        sql_rows = self._parse_sql_rows(payload.get("sql_result"))
        if sql_rows:
            data_sheet = workbook.create_sheet("Data")
            headers = list(sql_rows[0].keys())
            data_sheet.append(headers)
            for row in sql_rows:
                data_sheet.append([row.get(header, "") for header in headers])

        anomalies = payload.get("anomalies") or []
        if anomalies:
            anomaly_sheet = workbook.create_sheet("Anomalies")
            headers = ["field", "label", "value", "median", "delta", "direction"]
            anomaly_sheet.append(headers)
            for item in anomalies:
                anomaly_sheet.append([item.get(header, "") for header in headers])

        base_name = self._sanitize_filename(str(payload.get("question", "report")))
        file_path = self.export_dir / f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook.save(file_path)
        return file_path

    def export_pdf(self, payload: Dict[str, Any]) -> Path:
        base_name = self._sanitize_filename(str(payload.get("question", "report")))
        file_path = self.export_dir / f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        document = SimpleDocTemplate(
            str(file_path),
            pagesize=A4,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )
        styles = getSampleStyleSheet()
        normal = styles["BodyText"]
        normal.fontName = self.pdf_font_name
        normal.leading = 16
        title_style = styles["Title"]
        title_style.fontName = self.pdf_font_name
        heading = styles["Heading2"]
        heading.fontName = self.pdf_font_name

        story = [
            Paragraph("智能数据查询报告", title_style),
            Spacer(1, 8),
        ]

        for label, value in self._build_summary_rows(payload):
            story.append(Paragraph(f"<b>{label}：</b>{value}", normal))
            story.append(Spacer(1, 4))

        sources = payload.get("sources") or []
        if sources:
            story.append(Spacer(1, 6))
            story.append(Paragraph("参考来源", heading))
            for index, source in enumerate(sources[:10], start=1):
                story.append(Paragraph(f"{index}. {source}", normal))

        sql_rows = self._parse_sql_rows(payload.get("sql_result"))
        if sql_rows:
            story.append(Spacer(1, 8))
            story.append(Paragraph("数据样本", heading))
            headers = list(sql_rows[0].keys())
            table_data: List[List[str]] = [headers]
            for row in sql_rows[:10]:
                table_data.append([str(row.get(header, "")) for header in headers])
            table = Table(table_data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTNAME", (0, 0), (-1, -1), self.pdf_font_name),
                    ]
                )
            )
            story.append(table)

        document.build(story)
        return file_path

    def export(self, payload: Dict[str, Any], export_format: str) -> Dict[str, Any]:
        fmt = (export_format or "xlsx").lower()
        if fmt in {"xlsx", "excel"}:
            path = self.export_excel(payload)
            return {
                "path": str(path.resolve()),
                "filename": path.name,
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "format": "xlsx",
            }
        if fmt == "pdf":
            path = self.export_pdf(payload)
            return {
                "path": str(path.resolve()),
                "filename": path.name,
                "content_type": "application/pdf",
                "format": "pdf",
            }
        raise ValueError(f"不支持的导出格式: {export_format}")
